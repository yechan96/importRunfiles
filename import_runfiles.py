import os
import bs4 as bs
import argparse
import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment
from datetime import datetime
import glob
import sys

def importRunfile(file_path):
    try:
        # Read the content of the file
        with open(file_path, 'r') as file:
            content = file.read()

        # Create a BeautifulSoup object
        soup = bs.BeautifulSoup(content, 'html.parser')  # Use 'lxml' instead of 'html.parser' if you prefer lxml

        # Now, you can navigate and manipulate the parsed content using BeautifulSoup methods
        table = soup.find('table')
        if table is None:
            print(f"Error: No table found in the file. Please check the runfile: {file_path}.")
            sys.exit(1)
        table_rows = table.find_all('tr')
        parse_ID = {"Elution rack ID","File","Start time (yyyy-mm-dd hh:mm:ss)","QIAsymphony SP serial number","Software Version","Reagent rack description","User","Batch ID"}
        ls = {}
        sample_row = -2
        batch = []
        batch_num = -1
        curr_batch = 0
        samples  = []
        for tr in table_rows:
            td = tr.find_all('td')
            row = [i.text for i in td]
            if row[0] in parse_ID and row[0] not in ls:
                if row[0] != "User" and row[0] != "Batch ID":
                    ls[row[0]] = row[1]
                else:
                    ls[row[0]] = row[1:]
            
            if row[0] == "Samples":
                sample_row = -1
                samples = []
                batch_num += 1
            else: 
                if sample_row == -1:
                    sample_row = 0
                elif sample_row >= 0 and sample_row <= 23:
                    if " *"  in row[0]:
                        row[0] = row[0][:-3]
                    row[5] = row[5][0] + row[5][2:]
                    row[2] = int(row[2])
                    samples.append(row)
                    sample_row += 1

            if sample_row == 24 and curr_batch == batch_num:
                batch.append(samples)
                curr_batch += 1

        input_string = ls["Start time (yyyy-mm-dd hh:mm:ss)"]
        formatted_string = input_string.replace("&nbsp;", " ")
        date_object = datetime.strptime(formatted_string, "%Y-%m-%d %H:%M:%S")
        ls["Start time (yyyy-mm-dd hh:mm:ss)"] = date_object

        for i in range(len(batch)):
            for row in batch[i]:
                row += [ int(ls["Batch ID"][i]), ls["Reagent rack description"], ls["Elution rack ID"], ls["QIAsymphony SP serial number"], ls["Software Version"], ls["File"], ls["Start time (yyyy-mm-dd hh:mm:ss)"], ls["User"][i]]
        batchIndex = -1
        colCheck = batch[0][0][5]
        if colCheck == "A1":
            batchIndex = 0
        elif colCheck == "A4":
            batchIndex = 1
        elif colCheck == "A7":
            batchIndex = 2
        elif colCheck == "A10":
            batchIndex = 3

        return batchIndex, batch
    except FileNotFoundError:
        print(f"Error: File '{file_path}' not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


def append_row(sheet, data):
    # Determine the next available row
    next_row = sheet.max_row + 1

    # Iterate through the data and set values in each cell of the row
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=next_row, column=col_num, value=value)

def insert_row(sheet, row_index, data):
    # Shift existing rows down to make space for the new row
    sheet.insert_rows(row_index)

    # Set values in each cell of the new row
    for col_num, value in enumerate(data, start=1):
        sheet.cell(row=row_index, column=col_num, value=value)

def format_column_as_date(sheet, column_letter, start_row):
    # Convert the column letter to a column index
    column_index = column_index_from_string(column_letter)

    # Define a custom date style
    date_style = NamedStyle(name='date_style', number_format='mm/dd/yyyy')

    # Apply the custom date style to the specified column starting from the specified row
    for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            cell.style = date_style

def print_sheet_contents(sheet):
    # Print the contents of the sheet
    for row in sheet.iter_rows():
        for cell in row:
            print(f"{cell.coordinate}: {cell.value}")
        print()  # Add a newline between rows

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Import runfiles and generate a xlsx file.")
    parser.add_argument("-d", "--directory", required=True, help="Directory containing run files")
    parser.add_argument("-n", "--project_name", required=True, help="project name")
    parser.add_argument("-i", "--project_ID", required=True, help="project ID")

    args = parser.parse_args()

    # make an array for each of the batches
    master_batch = [None, None, None, None] 

    # get all runfiles
    run_files = glob.glob(args.directory+"/*"+args.project_ID+".htm*")

    if not run_files:
        print(f"Error: No files for {args.project_ID} found in the directory: {args.directory}")
        sys.exit(1)
    
    # main loop for gathering runfiles and attaching to master_batch
    for run_file in run_files:
        batchIndex, batches = importRunfile(run_file)

        for batch in batches:
            if master_batch[batchIndex] != None:
                print("Error: There are overlapping runfiles. Please check to see if the correct ones are in the directory.")
                sys.exit(1)
            master_batch[batchIndex] = batch
            batchIndex += 1

    # copy the template to generate a new excel sheet
    xlsx_fileName = args.project_name+" extraction import runfile "+args.project_ID+".xlsx"
   
    try:
        workbook = openpyxl.load_workbook("Extraction import template test.xlsx")
        sheet = workbook.worksheets[0]
        insert_row_index = 2

        for batch in master_batch:
            if batch == None:
                print("Error: There is a missing runfile. Please check to see if the correct ones are in the directory.")
                sys.exit(1)

        # append each row from the master_batch
        for batch in master_batch:
            for row in batch:
                insert_row(sheet, insert_row_index, row)
                insert_row_index += 1

        # Specify the column you want to format (e.g., column B) and the row after the header (e.g., row 2)
        column_letter_to_format = "S"
        start_row_after_header = 2

        # Format the column as "mm/dd/yyyy" after the header
        format_column_as_date(sheet, column_letter_to_format, start_row_after_header)

        # Save new workbook before align text to center and change DNA vol
        workbook.save(xlsx_fileName)
        
        # Center align all contents 
        workbook = openpyxl.load_workbook (xlsx_fileName)
        worksheet = workbook.active
        for row in range(1, worksheet.max_row+1):
            for col in range(1, worksheet.max_column+1):
                cell = worksheet.cell(row, col)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=False)

        # Replace DNA vol from 500 to post-quant/aliquot vol
        if "ATLAS" in args.project_name:
              for row in worksheet.iter_rows():
                  for cell in row:
                      if cell.value == "500 µl":
                          cell.value = "484"
        if "MO" in args.project_name:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.value == "500 µl":
                        cell.value = "488"
                        
        # Save finalized output file
        workbook.save(xlsx_fileName)
        print(f"Import successful!")
    except FileNotFoundError:
        print("Error: File 'Extraction import template test.xlsx` not found.")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

